import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import os
import sys
import pandas as pd
import random


def search_data(datalink):
    print("----------------------data search begin-----------------------------")
    response = requests.get(datalink, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.75 Safari/537.36',
    })
    # print(response.text)
    soup = BeautifulSoup(response.content, 'lxml')
    ktbody = soup.find_all('tbody')[2]
    ktr = ktbody.find_all('tr')

    # wb =openpyxl.Workbook()
    # sheet = wb['Sheet']
    wb = load_workbook('战舰战损数据汇总.xlsx')
    wb.guess_types = True
    sheet = wb.active
    # rows = []
    # for row in sheet.rows:
    #     for col in row:
    #         rows.append(col.value, end="\t")

    # for rowt in rows:
    #     print(rowt)
    # 创建excel文档第一行
    # sheet['A1'] = '击毁日期'
    # sheet['B1'] = '击毁时间'
    # sheet['C1'] = '价值'
    # sheet['D1'] = '具体损毁链接'
    # sheet['E1'] = '安全等级'
    # sheet['F1'] = '作战星系'
    # sheet['G1'] = '作战星域'
    # sheet['H1'] = '玩家ID'
    # sheet['I1'] = '玩家所属公司'
    # 创建样式（宽度）
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['D'].width = 25.0
    sheet.column_dimensions['G'].width = 15.0
    sheet.column_dimensions['H'].width = 10.0
    sheet.column_dimensions['I'].width = 35.0
    # 行高
    sheet.row_dimensions[1].height = 20
    i = 1
    for kth in ktr:
       # print("-----------------------------start---------------------------")
        kr = kth.find_all('th')  # 寻找日期
        if (kr):
            s_date = kr[0].get_text()

        ktd = kth.find_all('td')
        if (ktd):
            # print('击毁日期: ' + s_date)  # 数据来自上一个tr

            s_time = ktd[0].get_text()
            # print('击毁时间: ' + s_time[1:6])
            s_value = ktd[0].a.get_text()
            # print('价值：' + s_value)
            s_lostlink = 'https://zkillboard.com' + ktd[0].a['href']
            # print('具体损毁链接：' + s_lostlink)

            s_safelevel = ktd[2].find_all('span')[0].get_text()
            # print('安全等级：' + s_safelevel)
            s_location_xi = ktd[2].find_all('a')[0].get_text()
            # print('作战星系：' + s_location_xi)
            s_location_yu = ktd[2].find_all('a')[1].get_text()
            # print('作战星域：' + s_location_yu)

            s_playerId = ktd[4].find_all('a')[0].get_text()
            # print('玩家ID：' + s_playerId)
            s_playerCom = ktd[4].find_all('a')[1].get_text()
            # print('玩家所属公司：' + s_playerCom)
            s_boat_type = ktd[4].get_text().split('(')[1].split(')')[0]
            # print(s_boat_type)

            # 在添加数据之前，要进行数据唯一性检查，主要参考 日期+时间（和价值量）
            # first 组合数据
            # s = s_date+' '+s_time[1:6]+' '+s_value
            # 匹配,有问题
            # for rowt in rows:
            #     if s == rowt:  # if语句，不加括号！！！切记
            #         s1 = 1
            # if s1 != 1:
            sheet.append([s_date, s_time[1:6], s_boat_type, s_playerId, s_value,
                          s_location_xi, s_location_yu, s_safelevel, s_playerCom, s_lostlink])

        # l = []
        # for rowt in rows:
        #     if rowt not in l:
        #         l.append(x)
        # print(l)
        # # print(s)

        # sheet.append([s_date, s_time[1:6], s_value, s_lostlink, s_safelevel
        # , s_location_xi, s_location_yu, s_playerId, s_playerCom])

        # sheet['A'+str(i)] =s_date
        # sheet['B'+str(i)] =s_time[1:6]
        # sheet['C'+str(i)] =s_value
        # sheet['D'+str(i)] =s_lostlink
        # sheet['E'+str(i)] =s_safelevel
        # sheet['F'+str(i)] =s_location_xi
        # sheet['G'+str(i)] =s_location_yu
        # sheet['H'+str(i)] =s_playerId
        # sheet['I'+str(i)] =s_playerCom

        # 进行数据汇总
       # print('------------------------------end--------------------------')
        # if(i==10):
        #     break
        # else:
        i = i+1

    wb.save('战舰战损数据汇总.xlsx')
    print('爬完当前页')
    # 避免被拦截
    time.sleep(random.randint(10, 20))


def check_data(link):
    print('开始数据去重，数据处理起来比较复杂，请稍等.....')
    data = pd.DataFrame(pd.read_excel('战舰战损数据汇总.xlsx'))

    data.drop_duplicates(subset=None, keep='first',
                         inplace=True)  # data中一行元素全部相同时才去除
    # data.drop(data.columns[0], axis=1, inplace=True)
    # DataFrame.drop(labels=None,axis=0, index=None, columns=None, inplace=False)
    data.to_excel("战舰战损数据汇总.xlsx", index=False)


def find_excelfile():
    print('-------------check excel file -----')
    excelfile = os.path.exists('战舰战损数据汇总.xlsx')
    if excelfile == False:
        print('发现从未收集过此战舰的战损数据，\n现在程序已经创建“战舰战损数据汇总.xlsx，请别删除”')
        print('删除了，你损失就大了。。。，\n你又要重复爬虫和汇总很多次了')
        # 创建文件20190718 1907
        print('---------------creat excel file----')
        wb = Workbook()  # 创建文件对象
        # grab the active worksheet
        sheet = wb.active  # 获取第一个sheet
        # 创建excel文档第一行
        sheet['A1'] = '击毁日期'
        sheet['B1'] = '击毁时间'
        sheet['C1'] = '船型'
        sheet['D1'] = '玩家ID'
        sheet['E1'] = '价值'
        sheet['F1'] = '作战星系'
        sheet['G1'] = '作战星域'
        sheet['H1'] = '安全等级'
        sheet['I1'] = '玩家所属公司'
        sheet['J1'] = '具体损毁链接'

        wb.save('战舰战损数据汇总.xlsx')

    else:
        print('发现有存在有"战舰战损数据汇总.xlsx",进行下一步操作')

    print('check-done')


if __name__ == '__main__':
    try:
        print('--------------------------------------------------------------------------')
        print('   let\'s go !操作不可逆,还请谨慎操作!来,~~~~~让我们一起在eve里荡起双桨...')
        print('--------------------------------------------------------------------------')
        # 监测网站是否正常
        print('-------website checking-----------')
        rbase = requests.get('https://zkillboard.com',  headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.75 Safari/537.36',
        })
        if rbase.status_code != requests.codes.ok:
            raise Exception('网站(website): zkillboard.com 访问不正常,请核查')
        # 输入战舰相关信息
        else:
            print('website:https://zkillboard.com is Ok')
        print(
            '举例, 战舰: Machariel ,它的统计链接:https://zkillboard.com/ship/17738/losses/中的数字代码"17738"')
        datalinknumber = input('请输入战舰代码(数字):')
        if len(datalinknumber) != 5:
            raise Exception('战舰代码长度不对')
            # 检查数据合法性,抛出异常
        datalink = 'https://zkillboard.com/ship/'+datalinknumber+'/losses/'
        r = requests.get(datalink,  headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.75 Safari/537.36',
        })
        if r.status_code != requests.codes.ok:
            raise Exception('战舰代码有误,请求不到数据.请尝试在浏览器试试能否访问该战舰战损数据')

        print(
            '数据功能选择:   \n 1. 最近1页数据;   \n 2. 最近2-10页数据;  \n 3. 最近2-20页数据;  \n 4. 自定义读取链接(DIY) ')
        cantype = input("请输入功能代码(数字):")

        find_excelfile()
        # 寻找要存储的文件，并做提示
        if cantype == '1':  # 最近1页数据
            search_data(datalink)
        elif cantype == '2':
            m = 2
            while m <= 10:

                print('---------我是重复操作的分隔线 ，现在的时间是：' +
                      time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'--------------')
                # print('使用提醒：1. exe文件和xlsx表放在同一个文件夹内。 2,不能改统计表的名字。  3,输入链接时一定要确认战舰类型。')
                # datalink = input("输入要统计战舰损失的数据链接：")
                datalink = 'https://zkillboard.com/ship/'+datalinknumber+'/losses/page/' + \
                    str(m)+'/'
                search_data(datalink)
                print('执行完爬虫第  '+str(m)+'  页')
                m += 1
        elif cantype == '3':
            m = 2
            while m <= 20:

                print('---------我是重复操作的分隔线 ，现在的时间是：' +
                      time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'--------------')
                # print('使用提醒：1. exe文件和xlsx表放在同一个文件夹内。 2,不能改统计表的名字。  3,输入链接时一定要确认战舰类型。')
                # datalink = input("输入要统计战舰损失的数据链接：")
                datalink = 'https://zkillboard.com/ship/'+datalinknumber+'/losses/page/' + \
                    str(m)+'/'
                search_data(datalink)
                print('执行完爬虫第  '+str(m)+'  页')
                # check_data('a')
                # print('所有操作完成，请打开表格查看数据')
                m += 1
        elif cantype == '4':
            raise Exception('功能完善中......')
        else:
            raise Exception('选择错误,没有可操作的功能代码.')
        check_data('a')
        # m = 1
        # # 抓当前页数据
        # # 抓2-100的数据
        # while m <= 2:
        #     m += 1
        #     print('---------我是操作的分隔线 ，现在的时间是：' +
        #           time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'--------------')
        #     # print('使用提醒：1. exe文件和xlsx表放在同一个文件夹内。 2,不能改统计表的名字。  3,输入链接时一定要确认战舰类型。')
        #     # datalink = input("输入要统计战舰损失的数据链接：")
        #     # datalink = 'https://zkillboard.com/ship/17738/losses/page/' + \
        #     # str(m)+'/'
        #     search_data(datalink)
        #     print('执行完爬虫，准开始下一步')
        #     # check_data('a')
        #     # print('所有操作完成，请打开表格查看数据')
        #     time.sleep(5)
    except Exception as e:
        print('-----------------------------------------')
        print('操作错误! 详细信息:')
        print(e)

    finally:
        print('-----------------------------------------')
        print('DONE ,完成了.....该做的我都做了.我现在要自毁了,倒计时5秒开始,,,,')
        print('coder junjun '+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        print('二维码稍等就来....')
        time.sleep(5)
        sys.exit(0)
